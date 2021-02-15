import magic
import pytz
from io import BufferedReader
from openpyxl import Workbook, load_workbook
from datetime import date, datetime
from activities.models import *
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User, Group, Permission
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from .custom import *
from activities.forms import *
from django.urls import reverse
from django.core.files import File

#UPLOAD: 19
#BACKUP: 160
#RESTORE: 220

#UPLOAD ACTIVITY FROM SANITIZED EXCEL
def upload_webapp(ws_webapp):
	row_i=2
	while ws_webapp.cell(row=row_i, column=1).value != None:
		task = Task(
			assigned_by = get_object_or_404(UserProfile, pk=ws_webapp.cell(row=row_i, column=1).value),
		)
		task.save()
		new_application = ApplicationSecurity(
			task=task,
			name = ws_webapp.cell(row=row_i, column=2).value,
			url = ws_webapp.cell(row=row_i, column=3).value,
			category = ws_webapp.cell(row=row_i, column=5).value,
			accessibility= ws_webapp.cell(row=row_i, column=6).value,
			testing_type = ws_webapp.cell(row=row_i, column=7).value,
			environment = ws_webapp.cell(row=row_i, column=8).value,
			development = ws_webapp.cell(row=row_i, column=9).value,
			functionality = ws_webapp.cell(row=row_i, column=10).value,
			role_count = ws_webapp.cell(row=row_i, column=11).value,
			loc = ws_webapp.cell(row=row_i, column=12).value,
			owner =  ws_webapp.cell(row=row_i, column=13).value,
			spoc =  ws_webapp.cell(row=row_i, column=14).value,
			comments= ws_webapp.cell(row=row_i, column=15).value,
			critical = ws_webapp.cell(row=row_i, column=16).value if ws_webapp.cell(row=row_i, column=16).value != None else 'n',
		)
		new_application.save()
		row_i = row_i + 1

def upload_vapt(ws_vapt):
	row_i=2
	while ws_vapt.cell(row=row_i, column=1).value != None:
		task = Task(
			assigned_by = get_object_or_404(UserProfile, pk=ws_vapt.cell(row=row_i, column=1).value),
		)
		task.save()
		new_application = VaptAssessment(
			task=task,
			name= ws_vapt.cell(row=row_i, column=2).value,
			ip_address= ws_vapt.cell(row=row_i, column=3).value,
			item_count= ws_vapt.cell(row=row_i, column=4).value,
			accessibility= ws_vapt.cell(row=row_i, column=5).value,
			environment = ws_vapt.cell(row=row_i, column=6).value,
			device_type = ws_vapt.cell(row=row_i, column=7).value,
			owner =  ws_vapt.cell(row=row_i, column=8).value,
			spoc =  ws_vapt.cell(row=row_i, column=9).value,
			location =  ws_vapt.cell(row=row_i, column=10).value,
			comments= ws_vapt.cell(row=row_i, column=11).value,
		)
		new_application.save()
		row_i = row_i + 1

def upload_config(ws_config):
	row_i=2
	while ws_config.cell(row=row_i, column=1).value != None:
		task = Task(
			assigned_by = get_object_or_404(UserProfile, pk=ws_config.cell(row=row_i, column=1).value),
		)
		task.save()
		new_application = ConfigurationReview(
			task=task,
			name= ws_config.cell(row=row_i, column=2).value,
			host_name= ws_config.cell(row=row_i, column=3).value,
			item_count= ws_config.cell(row=row_i, column=4).value,
			device_type= ws_config.cell(row=row_i, column=5).value,
			owner =  ws_config.cell(row=row_i, column=6).value,
			spoc =  ws_config.cell(row=row_i, column=7).value,
			location =  ws_config.cell(row=row_i, column=8).value,
			comments= ws_config.cell(row=row_i, column=9).value,
		)
		new_application.save()
		row_i = row_i + 1

def add_data_to_db(excel):
	wb = load_workbook(excel[1:])
	if 'webapp' in wb.sheetnames:
		upload_webapp(wb['webapp'])		#20
	
	if 'vapt' in wb.sheetnames:
		upload_vapt(wb['vapt'])			#47
	
	if 'config' in wb.sheetnames:
		upload_config(wb['config'])		#70

@login_required
@group_required('Vendor')
def createActivity(request):
	if request.method == 'POST':
		form = UploadForm(request.POST, request.FILES)

		files = request.FILES['excel']
		file_type = magic.from_buffer(files.read())
		file_name = str(files)
		file_extension = file_name.split('.')
		valid_extension = ['xlsx', 'xls']

		if len(file_extension) > 2:
			context = {
                'message_type' : 'alert-danger',
		        'message_title' : 'Error!',
                'message_body' : 'File with multiple extension is not allowed!',
			}
		elif not file_extension[1] in valid_extension:
   			context = {
                'message_type' : 'alert-danger',
                'message_title' : 'Error!',
                'message_body' : 'File with given extension is not allowed!',
            }
		elif not "Microsoft Excel" in file_type:
			context = {
            	'message_type' : 'alert-danger',
            	'message_title' : 'Error!',
            	'message_body' : 'Invalid file type!',
            }
		elif form.is_valid():
			upload = SanitizedUpload(
				uploaded_by = get_object_or_404(UserProfile, pk=request.user.user_profile.pk),
				files = files,
			)
			upload.save()
			add_data_to_db(upload.files.url)		#91
			context = {
            	'message_type' : 'alert-success',
            	'message_title' : 'Success!',
            	'message_body' : 'Activities requested uploaded!',
            }
		else:
			context = {
           		'message_type' : 'alert-danger',
           		'message_title' : 'Error!',
           		'message_body' : 'Please check your input!',
        	}
	else:
		context = {}

	form = UploadForm()
	context.update({
       	'form': form,
    })

	return render(request, 'activities/upload.html', context=context)

#BACKUP
def add_data(sheet, data):
	if not data:
		return
	row_i = 1
	for row in data:
		sheet.cell(row=row_i, column=1, value=row.pk)
		col_i = 2
		row_arr = row.object2array()
		for cell in row_arr:
			sheet.cell(row=row_i, column=col_i, value=cell)
			col_i = col_i + 1
		row_i = row_i + 1

	return sheet

@login_required
@staff_member_required
def backup(request):
	backup_path = "../backup/"
	filename = str(date.today()) + "_-_Backup.xlsx"
	wb = Workbook()

	ws_organ = wb.active
	ws_organ.title = "organization"
	ws_depart = wb.create_sheet('department')
	ws_user = wb.create_sheet('user_profile')
	ws_task = wb.create_sheet('task')
	ws_app = wb.create_sheet('webapp')
	ws_vapt = wb.create_sheet('vapt')
	ws_config = wb.create_sheet('config')
	ws_bulk = wb.create_sheet('bulk')
	ws_sanitize = wb.create_sheet('sanitize')
	ws_backup = wb.create_sheet('backup')
	ws_restore = wb.create_sheet('restore')
	#not used
	ws_comment = wb.create_sheet('comment')
	ws_critical = wb.create_sheet('criticality')
	ws_vuln = wb.create_sheet('vulnerability')

	#Create Backup Entry
	BackupTracker(
		backedup_by = get_object_or_404(UserProfile, pk=request.user.user_profile.pk),
		path = backup_path + filename,
	).save()

	#Backing up
	ws_organ = add_data(ws_organ, Organization.objects.all())		#161
	ws_depart = add_data(ws_depart, Department.objects.all())
	ws_user = add_data(ws_user, UserProfile.objects.all())
	ws_task = add_data(ws_task, Task.objects.all())
	ws_app = add_data(ws_app, ApplicationSecurity.objects.all())
	ws_vapt = add_data(ws_vapt, VaptAssessment.objects.all())
	ws_config = add_data(ws_config, ConfigurationReview.objects.all())
	ws_bulk = add_data(ws_bulk, BulkActivity.objects.all())
	ws_sanitize = add_data(ws_sanitize, SanitizedUpload.objects.all())
	ws_backup = add_data(ws_backup, BackupTracker.objects.all())
	ws_restore = add_data(ws_restore, RestoreTracker.objects.all())
	#not used
	ws_comment = add_data(ws_comment, Comment.objects.all())
	ws_critical = add_data(ws_critical, Criticality.objects.all())
	ws_vuln = add_data(ws_vuln, Vulnerability.objects.all())

	#Save
	wb.save(backup_path + filename)

	return HttpResponseRedirect(reverse('home')) 

#RESTORE
def create_users_and_groups():
	test_pass = "p1a2s3s4w5"

	#Create IOCL Users
	iocl_admin = User.objects.create_user(username='iocl_admin', password=test_pass)
	iocl_cois = User.objects.create_user(username='iocl_cois', password=test_pass)
	iocl_mkho = User.objects.create_user(username='iocl_mkho', password=test_pass)
	iocl_rhq = User.objects.create_user(username='iocl_rhq', password=test_pass)
	iocl_plho = User.objects.create_user(username='iocl_plho', password=test_pass)
	iocl_rnd = User.objects.create_user(username='iocl_rnd', password=test_pass)
	iocl_bd = User.objects.create_user(username='iocl_bd', password=test_pass)

	#Create KPMG User
	kpmg_user1 = User.objects.create_user(username='kpmg_user1', password=test_pass)

	#Create Groups
	client, created = Group.objects.get_or_create(name='Client')
	vendor, created = Group.objects.get_or_create(name='Vendor')

	#Assign User to Group
	vendor.user_set.add(kpmg_user1)
	vendor.user_set.add(iocl_admin)
	client.user_set.add(iocl_cois)
	client.user_set.add(iocl_mkho)
	client.user_set.add(iocl_rhq)
	client.user_set.add(iocl_plho)
	client.user_set.add(iocl_rnd)
	client.user_set.add(iocl_bd)

def restore_organization(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		row_data = Organization(
			name = sheet.cell(row=row_i, column=2).value,
		)
		row_data.save()
		row_i = row_i + 1

def restore_department(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		organization = get_object_or_404(Organization, pk=sheet.cell(row=row_i, column=4).value)
		row_data = Department(
			name = sheet.cell(row=row_i, column=2).value,
			full_name = sheet.cell(row=row_i, column=3).value,
			organization = organization,
			location = sheet.cell(row=row_i, column=5).value,
			address = sheet.cell(row=row_i, column=6).value,
		)
		row_data.save()
		row_i = row_i + 1

def restore_user_profile(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		user = get_object_or_404(User, pk=sheet.cell(row=row_i, column=2).value)
		department = get_object_or_404(Department, pk=sheet.cell(row=row_i, column=5).value)
		row_data = UserProfile(
			user = user,
			employee_id = sheet.cell(row=row_i, column=3).value,
			contact = sheet.cell(row=row_i, column=4).value,
			department = department,
		)
		row_data.save()
		row_i = row_i + 1
	
def restore_task(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		assigned_by = get_object_or_404(UserProfile, pk=sheet.cell(row=row_i, column=3).value)
		value = sheet.cell(row=row_i, column=5).value
		assigned_to = get_object_or_404(UserProfile, pk=value) if value != None else None
		row_data = Task(
			assigned_by = assigned_by,
			date_posted = sheet.cell(row=row_i, column=4).value,
			assigned_to = assigned_to,
			status = sheet.cell(row=row_i, column=6).value,
		)
		row_data.save()
		row_i = row_i + 1
	
def restore_webapp(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		task = get_object_or_404(Task, pk=sheet.cell(row=row_i, column=2).value)
		row_data = ApplicationSecurity(
			task = task,
			name = sheet.cell(row=row_i, column=3).value,
			url = sheet.cell(row=row_i, column=4).value,
			item_count = sheet.cell(row=row_i, column=5).value,
			category = sheet.cell(row=row_i, column=6).value,
			accessibility = sheet.cell(row=row_i, column=7).value,
			testing_type = sheet.cell(row=row_i, column=8).value,
			environment = sheet.cell(row=row_i, column=9).value,
			development = sheet.cell(row=row_i, column=10).value,
			functionality = sheet.cell(row=row_i, column=11).value,
			role_count = sheet.cell(row=row_i, column=12).value,
			loc = sheet.cell(row=row_i, column=13).value,
			owner = sheet.cell(row=row_i, column=14).value,
			spoc = sheet.cell(row=row_i, column=15).value,
			comments = sheet.cell(row=row_i, column=16).value,
			critical = sheet.cell(row=row_i, column=17).value,
		)
		row_data.save()
		row_i = row_i + 1
	
def restore_vapt(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		task = get_object_or_404(Task, pk=sheet.cell(row=row_i, column=2).value)
		row_data = VaptAssessment(
			task = task,
			name = sheet.cell(row=row_i, column=3).value,
			ip_address = sheet.cell(row=row_i, column=4).value,
			item_count = sheet.cell(row=row_i, column=5).value,
			accessibility = sheet.cell(row=row_i, column=6).value,
			environment = sheet.cell(row=row_i, column=7).value,
			device_type = sheet.cell(row=row_i, column=8).value,
			owner = sheet.cell(row=row_i, column=9).value,
			spoc = sheet.cell(row=row_i, column=10).value,
			location = sheet.cell(row=row_i, column=11).value,
			comments = sheet.cell(row=row_i, column=12).value,
		)
		row_data.save()
		row_i = row_i + 1

def restore_config(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		task = get_object_or_404(Task, pk=sheet.cell(row=row_i, column=2).value)
		row_data = ConfigurationReview(
			task = task,
			name = sheet.cell(row=row_i, column=3).value,
			host_name = sheet.cell(row=row_i, column=4).value,
			item_count = sheet.cell(row=row_i, column=5).value,
			device_type = sheet.cell(row=row_i, column=6).value,
			owner = sheet.cell(row=row_i, column=7).value,
			spoc = sheet.cell(row=row_i, column=8).value,
			location = sheet.cell(row=row_i, column=9).value,
			comments = sheet.cell(row=row_i, column=10).value,
		)
		row_data.save()
		row_i = row_i + 1

def restore_bulk(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		task = get_object_or_404(Task, pk=sheet.cell(row=row_i, column=2).value)
		files = open(sheet.cell(row=row_i, column=4).value[1:], 'rb')
		row_data = BulkActivity(
			task = task,
			category = sheet.cell(row=row_i, column=3).value,
			files = File(files),
		)
		files.close()
		row_data.save()
		row_i = row_i + 1

def restore_sanitize(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		user = get_object_or_404(UserProfile, pk=sheet.cell(row=row_i, column=2).value)
		files = open(sheet.cell(row=row_i, column=4).value[1:], 'rb')
		row_data = SanitizedUpload(
			uploaded_by = user,
			date_uploaded = sheet.cell(row=row_i, column=3).value,
			files = File(files),
		)
		files.close()
		row_data.save()
		row_i = row_i + 1

def restore_backup(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		user = get_object_or_404(UserProfile, pk=sheet.cell(row=row_i, column=2).value)
		row_data = BackupTracker(
			backedup_by = user,
			date_backedup = sheet.cell(row=row_i, column=3).value,
			path = sheet.cell(row=row_i, column=4).value,
		)
		row_data.save()
		row_i = row_i + 1

def restore_restore(sheet):
	row_i=1
	while sheet.cell(row=row_i, column=1).value != None:
		user = get_object_or_404(UserProfile, pk=sheet.cell(row=row_i, column=2).value)
		files = open(sheet.cell(row=row_i, column=4).value[1:], 'rb')
		row_data = RestoreTracker(
			restored_by = user,
			date_restored = sheet.cell(row=row_i, column=3).value,
			files = File(files),
		)
		files.close()
		row_data.save()
		row_i = row_i + 1

def restore_comment(sheet):
	pass

def restore_critical(sheet):
	pass

def restore_vuln(sheet):
	pass

def restore_from(excel):
	wb = load_workbook(BufferedReader(excel))

	if 'organization' in wb.sheetnames:
		restore_organization(wb['organization'])			#251

	if 'department' in wb.sheetnames:
		restore_department(wb['department'])				#260

	if 'user_profile' in wb.sheetnames:
		restore_user_profile(wb['user_profile'])			#274

	if 'task' in wb.sheetnames:
		restore_task(wb['task'])							#288

	if 'webapp' in wb.sheetnames:
		restore_webapp(wb['webapp'])						#307

	if 'vapt' in wb.sheetnames:
		restore_vapt(wb['vapt'])							#332

	if 'config' in wb.sheetnames:
		restore_config(wb['config'])						#352
	
	if 'bulk' in wb.sheetnames:
		restore_bulk(wb['bulk'])							#370

	if 'sanitize' in wb.sheetnames:
		restore_sanitize(wb['sanitize'])

	if 'backup' in wb.sheetnames:
		restore_backup(wb['backup'])
	
	if 'restore' in wb.sheetnames:
		restore_restore(wb['restore'])

	if 'comment' in wb.sheetnames:
		restore_comment(wb['comment'])

	if 'critical' in wb.sheetnames:
		restore_critical(wb['critical'])

	if 'vuln' in wb.sheetnames:
		restore_vuln(wb['vuln'])

@login_required
@staff_member_required
def restore(request):
	if request.method == 'POST':
		form = UploadForm(request.POST, request.FILES)

		files = request.FILES['excel']
		file_type = magic.from_buffer(files.read())
		file_name = str(files)
		file_extension = file_name.split('.')
		valid_extension = ['xlsx', 'xls']

		if len(file_extension) > 2:
			context = {
                'message_type' : 'alert-danger',
		        'message_title' : 'Error!',
                'message_body' : 'File with multiple extension is not allowed!',
			}
		elif not file_extension[1] in valid_extension:
   			context = {
                'message_type' : 'alert-danger',
                'message_title' : 'Error!',
                'message_body' : 'File with given extension is not allowed!',
            }
		elif not "Microsoft Excel" in file_type:
			context = {
            	'message_type' : 'alert-danger',
            	'message_title' : 'Error!',
            	'message_body' : 'Invalid file type!',
            }
		elif form.is_valid():
			restored_by = get_object_or_404(User, pk=request.user.pk)
			rst = RestoreTracker(
				restored_by = restored_by,
				files = files,
			)
			rst.save()
			create_users_and_groups()			#221
			restore_from(files)					#251
			context = {
            	'message_type' : 'alert-success',
            	'message_title' : 'Success!',
            	'message_body' : 'Activities requested uploaded!',
            }
		else:
			context = {
           		'message_type' : 'alert-danger',
           		'message_title' : 'Error!',
           		'message_body' : 'Please check your input!',
        	}
	else:
		context = {}

	form = UploadForm()
	context.update({
       	'form': form,
    })

	return render(request, 'activities/upload.html', context=context)
